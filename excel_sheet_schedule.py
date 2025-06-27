import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Load CSV
csv_path = "final_workshop_fixed.csv"  # adjust as needed
df = pd.read_csv(csv_path)

# Initialize Excel workbook
wb = Workbook()
wb.remove(wb.active)

# Styles
header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
bold_font = Font(bold=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def clean_sheet_name(day, session):
    session_label = {0: "All-day", 1: "Morning", 2: "Afternoon"}.get(session, f"Session{session}")
    return f"{day} {session_label}"

# Generate one sheet per session
unique_sessions = df[['Day', 'Session']].drop_duplicates()

for _, row in unique_sessions.iterrows():
    day, session = row['Day'], row['Session']
    sheet_name = clean_sheet_name(day, session)
    ws = wb.create_sheet(title=sheet_name[:31])

    # Filter relevant data
    session_data = df[(df['Day'] == day) & (df['Session'] == session)]
    grouped = session_data.groupby('Workshop Title')['Student'].apply(list).reset_index()

    # Store column widths
    col_widths = {}

    for col_idx, (_, grp) in enumerate(grouped.iterrows(), start=1):
        workshop = grp['Workshop Title']
        students = grp['Student']
        col_letter = get_column_letter(col_idx)

        # Header cell
        header_cell = ws.cell(row=1, column=col_idx, value=workshop)
        header_cell.font = bold_font
        header_cell.fill = header_fill
        header_cell.alignment = Alignment(horizontal='center')
        header_cell.border = thin_border
        col_widths[col_letter] = len(str(workshop))

        # Student cells
        for row_idx, student in enumerate(students, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=student)
            cell.border = thin_border
            col_widths[col_letter] = max(col_widths[col_letter], len(str(student)))

    # Set column widths
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width + 2  # Add padding

# Save the Excel file
output_path = "final_workshop_schedule_pretty.xlsx"
wb.save(output_path)
print(f"Saved styled Excel to: {output_path}")
